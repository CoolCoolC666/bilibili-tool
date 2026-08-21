"""v3.1.0+ XLSX 整合分析工具。

把 `output/xlsx/author_detail_*.xlsx` 多个文件合并分析，输出：
  1. 基础统计（条数、覆盖率、时间跨度）
  2. 年度分布矩阵
  3. 爆款视频 TOP 10
  4. 互动率（赞/播放 / 币/播放）
  5. 抓取时间分布
  6. 抓取质量（空标题 / 不完整数据）
  7. 有趣发现（最高币赞比 / 最长视频 / 短小精悍）
  8. 匿名化映射表

v3.1.1 隐私：UP 主姓名全部匿名化为 `UP-A/B/C/D`（按 UID 升序）。
真实 UP 名仅在脚本内部 _REAL_NAMES 字典中，不写入输出文件。

用法：
    # 默认：读 output/xlsx/ 目录
    py -3.13 tests/analyze_xlsx.py

    # 自定义目录
    py -3.13 tests/analyze_xlsx.py --data-dir "E:/桌面/.../整合"

    # 输出到指定文件
    py -3.13 tests/analyze_xlsx.py --out _analyze_out.txt
"""
from __future__ import annotations

import argparse
import io
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime

# 强制 UTF-8 输出（避免 Windows GBK 终端乱码）
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook  # noqa: E402


# ============================================================================
# 匿名化配置（v3.1.1 隐私）
# ============================================================================

# 真实 UP 名：仅在脚本内部用于标题脱敏，**不写入输出文件**
_REAL_NAMES = {
    53456: "Warma",
    131692216: "饭团Fine",
    399959326: "国服第一三月七",
    1751577265: "春山响hibiki",
}

# 公开标签：按 UID 升序 A → D
ANON_MAP = {
    53456: "UP-A",
    131692216: "UP-B",
    399959326: "UP-C",
    1751577265: "UP-D",
}

# 标题脱敏：把 real name 字符串从 title 里替换为公开标签
_TITLE_REPLACE = {
    "warma": "UP-A", "WARMA": "UP-A", "Warma": "UP-A",
    "沃玛": "UP-A",
    "春山响": "UP-D", "hibiki": "UP-D", "Hibiki": "UP-D", "HIBIKI": "UP-D",
    "饭团Fine": "UP-B", "饭团": "UP-B", "Fine": "UP-B",
    "国服第一三月七": "UP-C", "三月七": "UP-C",
}

# 真实投稿数（uapis 报告的 total，按 v3.0.5 诊断）
KNOWN_TOTALS = {
    1751577265: 983,   # 春山响
    53456: 262,        # Warma
}


# ============================================================================
# 核心分析函数
# ============================================================================

def parse_dt(s):
    if not s:
        return None
    try:
        return datetime.strptime(str(s)[:19], "%Y-%m-%d %H:%M:%S")
    except (ValueError, TypeError):
        return None


def load_xlsx_files(data_dir: str) -> dict:
    """从 data_dir 加载所有 author_detail_*.xlsx，返回 {uid: [record]}。"""
    all_data: dict = {}
    if not os.path.isdir(data_dir):
        print(f"❌ 目录不存在：{data_dir}", file=sys.stderr)
        return all_data

    for name in sorted(os.listdir(data_dir)):
        if not name.endswith(".xlsx"):
            continue
        if not name.startswith("author_detail_"):
            continue
        full = os.path.join(data_dir, name)
        wb = load_workbook(full, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        idx = {h: i for i, h in enumerate(header)}
        # 从文件名提取 uid
        parts = name.replace(".xlsx", "").split("_")
        try:
            uid = int(parts[2])
        except (ValueError, IndexError):
            continue
        records = []
        for row in rows[1:]:
            try:
                raw_title = row[idx["标题"]] or ""
                # v3.1.1 标题脱敏
                anon_title = raw_title
                for old, new in _TITLE_REPLACE.items():
                    if old in anon_title:
                        anon_title = anon_title.replace(old, new)
                records.append({
                    "uid": uid,
                    "bvid": row[idx["BV号"]],
                    "title": anon_title,
                    "up": ANON_MAP.get(uid, f"UP-?{uid}"),
                    "play": int(row[idx["播放量"]] or 0),
                    "like": int(row[idx["点赞"]] or 0),
                    "coin": int(row[idx["投币"]] or 0),
                    "favorite": int(row[idx["收藏"]] or 0),
                    "share": int(row[idx["分享"]] or 0),
                    "reply": int(row[idx["评论"]] or 0),
                    "danmaku": int(row[idx["弹幕"]] or 0),
                    "duration": int(row[idx["时长(秒)"]] or 0),
                    "pubdate": parse_dt(row[idx["发布时间"]]),
                })
            except (ValueError, KeyError, TypeError):
                continue
        all_data[uid] = records
    return all_data


def render_report(all_data: dict, data_dir: str) -> str:
    """渲染分析报告为字符串。"""
    out = []
    out.append("=" * 70)
    out.append(f"📊 XLSX 整合分析（v3.1.0+ analyze_xlsx.py · 目录：{data_dir}）")
    out.append("=" * 70)

    if not all_data:
        out.append("\n❌ 没有找到 author_detail_*.xlsx 文件")
        out.append("提示：先跑 Web 抓取，让 author_detail_*.xlsx 生成在 output/xlsx/ 目录")
        return "\n".join(out)

    # 1. 基础统计
    out.append("\n## 1. 基础统计")
    total_videos = 0
    for uid, records in sorted(all_data.items()):
        real_total = KNOWN_TOTALS.get(uid)
        pct = (len(records) / real_total * 100) if real_total else None
        up_name = records[0]["up"] if records else "?"
        pct_str = f"  (覆盖率 {pct:.1f}%, 真实 {real_total})" if pct else ""
        dates = [r["pubdate"] for r in records if r["pubdate"]]
        date_range = ""
        if dates:
            date_range = f"  [{min(dates).strftime('%Y-%m-%d')} ~ {max(dates).strftime('%Y-%m-%d')}]"
        out.append(f"  {up_name:<6} (UID {uid:>10}) {len(records):>4} 条{pct_str}{date_range}")
        total_videos += len(records)
    out.append(f"  {'合计':<35} {total_videos:>4} 条")

    # 2. 年度分布
    out.append("\n## 2. 年度分布（每个 UP 每年发布多少条）")
    year_by_up = defaultdict(Counter)
    for uid, records in all_data.items():
        for r in records:
            if r["pubdate"]:
                year_by_up[uid][r["pubdate"].year] += 1
    all_years = sorted(set(y for c in year_by_up.values() for y in c))
    header = "  UP   " + " ".join(f"{y:>5}" for y in all_years) + "  合计"
    out.append(header)
    for uid, records in sorted(all_data.items()):
        up_name = records[0]["up"] if records else "?"
        row_counts = [year_by_up[uid].get(y, 0) for y in all_years]
        out.append(f"  {up_name:<5}" + " ".join(f"{c:>5}" for c in row_counts) + f"  {sum(row_counts):>5}")

    # 3. 爆款 TOP 10
    out.append("\n## 3. 爆款视频 TOP 10（按播放量）")
    all_records = []
    for records in all_data.values():
        all_records.extend(records)
    top10 = sorted(all_records, key=lambda r: r["play"], reverse=True)[:10]
    for i, r in enumerate(top10, 1):
        out.append(f"  {i:>2}. [{r['up']:<5}] {r['play']:>9,} 播放  {r['like']:>6,} 赞  {r['coin']:>5,} 币  {r['favorite']:>5,} 收")
        out.append(f"        {r['title'][:60]}")
        out.append(f"        {r['bvid']}  {r['pubdate'].strftime('%Y-%m-%d') if r['pubdate'] else '?'}  时长 {r['duration']}s")

    # 4. 互动率
    out.append("\n## 4. 互动率（赞/播放 ×100%）")
    for uid, records in sorted(all_data.items()):
        up_name = records[0]["up"] if records else "?"
        total_play = sum(r["play"] for r in records)
        total_like = sum(r["like"] for r in records)
        total_coin = sum(r["coin"] for r in records)
        total_fav = sum(r["favorite"] for r in records)
        if total_play == 0:
            continue
        out.append(f"  {up_name:<5} 播放 {total_play:>10,}  赞 {total_like:>7,}({total_like/total_play*100:>5.2f}%)  币 {total_coin:>6,}({total_coin/total_play*100:.2f}%)  收 {total_fav:>6,}({total_fav/total_play*100:.2f}%)")

    # 5. 抓取时间
    out.append("\n## 5. 抓取时间分布（看抓取速度）")
    times = []
    for name in sorted(os.listdir(data_dir)):
        if name.endswith(".xlsx") and name.startswith("author_detail_"):
            t = os.path.getmtime(os.path.join(data_dir, name))
            times.append((datetime.fromtimestamp(t), name))
    times.sort()
    for t, name in times:
        out.append(f"  {t.strftime('%Y-%m-%d %H:%M:%S')}  {name}")

    # 6. 质量
    out.append("\n## 6. 抓取质量")
    for uid, records in sorted(all_data.items()):
        up_name = records[0]["up"] if records else "?"
        ok_count = sum(1 for r in records if r["bvid"] and r["title"])
        no_title = sum(1 for r in records if not r["title"])
        out.append(f"  {up_name:<5} {ok_count}/{len(records)} 有完整数据 (空标题 {no_title})")

    # 7. 有趣发现
    out.append("\n## 7. 有趣发现")
    with_ratio = [
        (r, r["coin"] / r["like"] if r["like"] > 0 else 0)
        for r in all_records if r["like"] > 100
    ]
    top_ratio = sorted(with_ratio, key=lambda x: x[1], reverse=True)[:3]
    out.append("  7.1 最高「币/赞」比（前 3）:")
    for r, ratio in top_ratio:
        out.append(f"    [{r['up']:<5}] 币 {r['coin']:>5,} 赞 {r['like']:>6,}  比 {ratio:.2f}  - {r['title'][:40]}")
    longest = sorted(all_records, key=lambda r: r["duration"], reverse=True)[:3]
    out.append("  7.2 最长视频（前 3）:")
    for r in longest:
        mins = r["duration"] // 60
        secs = r["duration"] % 60
        out.append(f"    [{r['up']:<5}] {mins}m{secs}s  {r['play']:>8,} 播放  - {r['title'][:40]}")
    short_pop = [r for r in all_records if r["duration"] > 0 and r["duration"] < 120 and r["play"] > 100000]
    out.append(f"  7.3 短小精悍（< 2 分钟 + 播放 > 10 万）: {len(short_pop)} 条")
    for r in short_pop[:3]:
        out.append(f"    [{r['up']:<5}] {r['duration']}s  {r['play']:>8,} 播放  - {r['title'][:40]}")

    # 8. 匿名化映射表
    out.append("\n" + "=" * 70)
    out.append("## 8. 匿名化映射（公开版）")
    out.append("v3.1.1 隐私：以下只显示公开标签 + UID。**真实 UP 名仅在脚本内部，")
    out.append("不写入输出文件、不写入 README、不写入公开报告。**")
    out.append("需要回查时按 UID 查 B 站即可。")
    out.append("")
    out.append("| 公开标签 | UID |")
    out.append("|---|---|")
    for uid in sorted(ANON_MAP.keys()):
        out.append(f"| {ANON_MAP[uid]} | {uid} |")
    out.append("")
    out.append("排序说明：按 UID 升序 A → D，便于脚本内单字母对比")

    return "\n".join(out)


# ============================================================================
# CLI
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="v3.1.0+ XLSX 整合分析工具（按 UP 主匿名化）",
    )
    here = os.path.dirname(os.path.abspath(__file__))
    default_data = os.path.join(os.path.dirname(here), "output", "xlsx")
    parser.add_argument(
        "--data-dir", default=default_data,
        help=f"XLSX 所在目录（默认：{default_data}）",
    )
    parser.add_argument(
        "--out", default=None,
        help="输出文件路径（默认：stdout）",
    )
    args = parser.parse_args()

    print(f"📂 读取目录：{args.data_dir}", file=sys.stderr)
    all_data = load_xlsx_files(args.data_dir)
    if all_data:
        print(f"✅ 加载 {len(all_data)} 个 UP 主，{sum(len(r) for r in all_data.values())} 条视频",
              file=sys.stderr)
    report = render_report(all_data, args.data_dir)

    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(report)
        print(f"💾 已写入：{args.out}", file=sys.stderr)
    else:
        print(report)


if __name__ == "__main__":
    main()
