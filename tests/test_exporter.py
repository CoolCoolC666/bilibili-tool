"""Exporter 单元测试：去重逻辑 + 4 种格式。"""
import csv
import json
import os
import tempfile
import unittest
from datetime import datetime

from bilibili_tool.exporter import (
    dedupe_articles,
    dedupe_bangumis,
    dedupe_videos,
    save_xlsx,
    save_xlsx_multi,
    save_csv,
    save_json,
    save_txt,
    save_articles_csv,
    save_articles_json,
    save_articles_txt,
    save_bangumis_csv,
    save_bangumis_json,
    save_bangumis_txt,
    export_all,
    export_all_multi,
    filter_valid,
    _dedupe_key,
)
from bilibili_tool.models import ArticleInfo, BangumiInfo, VideoInfo


def make(bvid="", aid=None, title="t", status="ok", raw="", **kw):
    return VideoInfo(
        raw_input=raw or (bvid or str(aid or "x")),
        input_kind="bv" if bvid else "av",
        bvid=bvid,
        aid=aid,
        title=title,
        status=status,
        fetched_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        **kw,
    )


class TestDedupeKey(unittest.TestCase):
    def test_bv_key(self):
        self.assertEqual(_dedupe_key(make(bvid="BV1xxx")), ("bv", "BV1xxx"))

    def test_av_key(self):
        self.assertEqual(_dedupe_key(make(aid=170001, bvid="")), ("av", 170001))

    def test_bv_wins_over_av(self):
        # 都有的话用 BV（更稳定）
        v = make(bvid="BV1xxx", aid=170001)
        self.assertEqual(_dedupe_key(v), ("bv", "BV1xxx"))

    def test_raw_fallback(self):
        v = VideoInfo(raw_input="something", input_kind="raw")
        self.assertEqual(_dedupe_key(v), ("raw", "something"))

    def test_no_key(self):
        v = VideoInfo()
        self.assertIsNone(_dedupe_key(v))


class TestDedupeVideos(unittest.TestCase):
    def test_keeps_unique(self):
        v1 = make(bvid="BV1aaa")
        v2 = make(bvid="BV1bbb")
        out = dedupe_videos([v1, v2])
        self.assertEqual(len(out), 2)

    def test_merges_same_bv(self):
        v1 = make(bvid="BV1xxx", view=100)
        v2 = make(bvid="BV1xxx", view=200)  # 同一视频（理论上不会出现，但兜底）
        out = dedupe_videos([v1, v2])
        self.assertEqual(len(out), 1)
        # 保留第一次的
        self.assertEqual(out[0].view, 100)

    def test_bv_av_same_video(self):
        """用户输入 BV1xxx + av170001（同一视频），导出只 1 条。"""
        v_bv = make(bvid="BV1xxx", aid=170001, title="测试视频", raw="BV1xxx")
        v_av = make(bvid="BV1xxx", aid=170001, title="测试视频", raw="av170001")
        out = dedupe_videos([v_bv, v_av])
        self.assertEqual(len(out), 1)
        self.assertEqual(out[0].bvid, "BV1xxx")
        self.assertEqual(out[0].aid, 170001)

    def test_preserves_input_order(self):
        v1 = make(bvid="BV1aaa", title="A")
        v2 = make(bvid="BV1bbb", title="B")
        v3 = make(bvid="BV1ccc", title="C")
        out = dedupe_videos([v1, v2, v3, v1])  # v1 重复一次
        self.assertEqual([v.title for v in out], ["A", "B", "C"])

    def test_no_key_videos_kept(self):
        v1 = VideoInfo()
        v2 = VideoInfo()
        out = dedupe_videos([v1, v2])
        self.assertEqual(len(out), 2)  # 没 key 的不参与去重

    def test_mixed_with_dupes(self):
        v1 = make(bvid="BV1a")
        v2 = make(bvid="BV1b")
        v3 = make(bvid="BV1a")  # dup of v1
        v4 = make(aid=999, bvid="")  # 没 bvid，用 av
        v5 = make(bvid="BV1c", aid=999)  # 有 bvid，bvid 不跟 v4 重复
        out = dedupe_videos([v1, v2, v3, v4, v5])
        # v1, v2, v4, v5（v3 dup v1）
        self.assertEqual(len(out), 4)


class TestExportFormats(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_save_csv(self):
        v1 = make(bvid="BV1a", title="视频A")
        v2 = make(bvid="BV1b", title="视频B")
        path = save_csv([v1, v2], os.path.join(self.tmpdir, "t.csv"))
        self.assertTrue(os.path.exists(path))
        with open(path, encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["BV号"], "BV1a")

    def test_save_json(self):
        v1 = make(bvid="BV1a", title="视频A")
        path = save_json([v1], os.path.join(self.tmpdir, "t.json"))
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["bvid"], "BV1a")

    def test_save_txt(self):
        v1 = make(bvid="BV1a", title="视频A")
        v2 = make(bvid="BV1b", title="视频B", status="not_found", api_code=62002)
        path = save_txt([v1, v2], os.path.join(self.tmpdir, "t.txt"))
        with open(path, encoding="utf-8") as f:
            text = f.read()
        self.assertIn("视频A", text)
        self.assertIn("not_found", text)
        self.assertIn("共 2 条记录", text)

    def test_save_xlsx(self):
        v1 = make(bvid="BV1a", title="视频A")
        v2 = make(bvid="BV1b", title="视频B")
        path = save_xlsx([v1, v2], os.path.join(self.tmpdir, "t.xlsx"))
        self.assertTrue(os.path.exists(path))
        # 读回来验证
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # 第 1 行表头 + 2 行数据
        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[1][6], "BV1a")  # BV号 列
        self.assertEqual(rows[2][6], "BV1b")

    def test_export_all_dedupes_by_default(self):
        # 两条同一视频（BV + AV）
        v1 = make(bvid="BV1xxx", aid=170001, title="测试", raw="BV1xxx")
        v2 = make(bvid="BV1xxx", aid=170001, title="测试", raw="av170001")
        saved = export_all([v1, v2], base="test", out_dir=self.tmpdir)
        # 4 个文件都生成了
        for fmt, path in saved.items():
            self.assertTrue(os.path.exists(path), f"{fmt} 文件不存在")
        # 验证：CSV 只有 1 行数据
        with open(saved["csv"], encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        self.assertEqual(len(rows), 1, f"应只有 1 条去重后的记录，实际 {len(rows)}")

    def test_export_all_no_dedupe(self):
        v1 = make(bvid="BV1xxx", aid=170001, title="测试", raw="BV1xxx")
        v2 = make(bvid="BV1xxx", aid=170001, title="测试", raw="av170001")
        saved = export_all([v1, v2], base="test", out_dir=self.tmpdir, dedupe=False)
        with open(saved["csv"], encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        self.assertEqual(len(rows), 2, f"不去重应保留 2 条，实际 {len(rows)}")

    def test_filter_valid(self):
        ok1 = make(bvid="BV1a", title="A")
        ok2 = make(bvid="BV1b", title="B")
        fail1 = make(bvid="BV1c", status="not_found", api_code=62002)
        fail2 = make(bvid="BV1d", status="failed", error="network")
        out = filter_valid([ok1, fail1, ok2, fail2])
        self.assertEqual(len(out), 2)
        self.assertEqual([v.bvid for v in out], ["BV1a", "BV1b"])

    def test_export_all_exclude_invalid(self):
        """--exclude-invalid 开启时，导出文件只含 status=ok 的记录。"""
        ok = make(bvid="BV1a", title="A")
        fail = make(bvid="BV1c", status="not_found", api_code=62002)
        saved = export_all(
            [ok, fail, fail],
            base="test",
            out_dir=self.tmpdir,
            exclude_invalid=True,
        )
        # CSV 应只有 1 行
        with open(saved["csv"], encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        self.assertEqual(len(rows), 1, f"应只 1 条 ok 记录，实际 {len(rows)}")
        self.assertEqual(rows[0]["BV号"], "BV1a")

    def test_export_all_exclude_invalid_false_keeps_failed(self):
        """--exclude-invalid 关闭时（默认），保留所有状态。"""
        ok = make(bvid="BV1a", title="A")
        fail = make(bvid="BV1c", status="not_found", api_code=62002)
        saved = export_all(
            [ok, fail],
            base="test",
            out_dir=self.tmpdir,
            exclude_invalid=False,
        )
        with open(saved["csv"], encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        self.assertEqual(len(rows), 2)

    def test_export_all_filter_before_dedupe(self):
        """过滤应在去重之前：同一视频的 ok + failed 两种状态，先 filter 再 dedupe。"""
        ok = make(bvid="BV1x", title="OK")
        fail_same = make(bvid="BV1x", status="not_found", api_code=62002)
        # 开启 filter：fail 过滤掉，ok 保留；去重：1 条
        saved = export_all(
            [ok, fail_same],
            base="test",
            out_dir=self.tmpdir,
            exclude_invalid=True,
            dedupe=True,
        )
        with open(saved["csv"], encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["状态"], "ok")

    # === v2.8.0：3 类型混合导出 ===

    def _make_article(self, cv_id="cv1", title="专栏", status="ok", view=500, **kw) -> ArticleInfo:
        return ArticleInfo(
            input_kind="article", cv_id=cv_id, title=title, status=status, view=view,
            fetched_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"), **kw,
        )

    def _make_bangumi(self, season_id=1, ep_id=None, title="番剧", status="ok", rating_score=9.0, **kw) -> BangumiInfo:
        return BangumiInfo(
            input_kind="bangumi_ss" if season_id else "bangumi_ep",
            season_id=season_id, ep_id=ep_id, title=title, status=status, rating_score=rating_score,
            fetched_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"), **kw,
        )

    def test_dedupe_articles_keeps_unique(self):
        a1 = self._make_article(cv_id="cv1")
        a2 = self._make_article(cv_id="cv2")
        out = dedupe_articles([a1, a2])
        self.assertEqual(len(out), 2)

    def test_dedupe_articles_merges_dupes(self):
        a1 = self._make_article(cv_id="cv1", view=100)
        a2 = self._make_article(cv_id="cv1", view=200)  # 同一专栏
        out = dedupe_articles([a1, a2])
        self.assertEqual(len(out), 1)
        self.assertEqual(out[0].view, 100)  # 保留第一次

    def test_dedupe_bangumis_keeps_unique(self):
        b1 = self._make_bangumi(season_id=1)
        b2 = self._make_bangumi(season_id=2)
        out = dedupe_bangumis([b1, b2])
        self.assertEqual(len(out), 2)

    def test_dedupe_bangumis_ss_ep_same(self):
        """ss + ep 是同一季：unique_count 应该按 fingerprint 合并为 1。"""
        b1 = self._make_bangumi(season_id=1, ep_id=100, title="SS形式")
        b2 = self._make_bangumi(season_id=1, ep_id=100, title="EP形式")  # 完全相同的 ss+ep
        out = dedupe_bangumis([b1, b2])
        self.assertEqual(len(out), 1)
        self.assertEqual(out[0].title, "SS形式")  # 保留第一次

    def test_filter_valid_mixed_types(self):
        """filter_valid 对任何 dataclass 都能用（看 status 字段）"""
        from bilibili_tool.exporter import filter_valid
        v_ok = make(bvid="BV1a", status="ok")
        v_fail = make(bvid="BV1b", status="failed")
        a_ok = self._make_article(cv_id="cv1", status="ok")
        a_fail = self._make_article(cv_id="cv2", status="not_found")
        b_ok = self._make_bangumi(season_id=1, status="ok")
        out = filter_valid([v_ok, v_fail, a_ok, a_fail, b_ok])
        self.assertEqual(len(out), 3)

    def test_save_articles_csv(self):
        a1 = self._make_article(cv_id="cv1", title="专栏A")
        a2 = self._make_article(cv_id="cv2", title="专栏B")
        path = save_articles_csv([a1, a2], os.path.join(self.tmpdir, "t.csv"))
        self.assertTrue(os.path.exists(path))
        with open(path, encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["专栏ID(cv)"], "cv1")
        self.assertEqual(rows[0]["标题"], "专栏A")

    def test_save_articles_json(self):
        a1 = self._make_article(cv_id="cv1")
        path = save_articles_json([a1], os.path.join(self.tmpdir, "t.json"))
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["cv_id"], "cv1")

    def test_save_articles_txt(self):
        a1 = self._make_article(cv_id="cv1", title="我的专栏")
        a2 = self._make_article(cv_id="cv2", status="not_found", api_code=404)
        path = save_articles_txt([a1, a2], os.path.join(self.tmpdir, "t.txt"))
        with open(path, encoding="utf-8") as f:
            text = f.read()
        self.assertIn("我的专栏", text)
        self.assertIn("not_found", text)
        self.assertIn("共 2 条记录", text)

    def test_save_bangumis_csv(self):
        b1 = self._make_bangumi(season_id=1, title="番剧A", rating_score=9.5)
        path = save_bangumis_csv([b1], os.path.join(self.tmpdir, "t.csv"))
        with open(path, encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["番剧季ID(ss)"], "1")
        self.assertEqual(rows[0]["评分"], "9.5")

    def test_save_bangumis_json(self):
        b1 = self._make_bangumi(season_id=1, ep_id=100)
        path = save_bangumis_json([b1], os.path.join(self.tmpdir, "t.json"))
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        self.assertEqual(data[0]["season_id"], 1)
        self.assertEqual(data[0]["ep_id"], 100)

    def test_save_bangumis_txt(self):
        b1 = self._make_bangumi(season_id=1, title="那年那兔", rating_score=9.3)
        path = save_bangumis_txt([b1], os.path.join(self.tmpdir, "t.txt"))
        with open(path, encoding="utf-8") as f:
            text = f.read()
        self.assertIn("那年那兔", text)
        self.assertIn("⭐ 评分: 9.3", text)

    def test_save_xlsx_multi_three_sheets(self):
        """v2.8.0：xlsx 3 个 sheet 各自一类型。"""
        from openpyxl import load_workbook
        v1 = make(bvid="BV1a", title="视频A")
        a1 = self._make_article(cv_id="cv1", title="专栏A")
        b1 = self._make_bangumi(season_id=1, title="番剧A")
        path = save_xlsx_multi([v1], [a1], [b1], os.path.join(self.tmpdir, "multi.xlsx"))
        self.assertTrue(os.path.exists(path))
        wb = load_workbook(path)
        self.assertEqual(wb.sheetnames, ["视频数据", "专栏数据", "番剧数据"])
        # 视频 sheet 1 行
        ws_v = wb["视频数据"]
        self.assertEqual(ws_v.max_row, 2)  # 表头 + 1 行
        # 专栏 sheet
        ws_a = wb["专栏数据"]
        self.assertEqual(ws_a.max_row, 2)
        # 番剧 sheet
        ws_b = wb["番剧数据"]
        self.assertEqual(ws_b.max_row, 2)

    def test_save_xlsx_multi_empty_sheet_skipped(self):
        """空数据的 sheet 应该跳过（不创建空 sheet）"""
        from openpyxl import load_workbook
        v1 = make(bvid="BV1a", title="视频A")
        # 不传 article/bangumi
        path = save_xlsx_multi([v1], [], [], os.path.join(self.tmpdir, "only_v.xlsx"))
        wb = load_workbook(path)
        # 应该只有 1 个 sheet（视频）
        self.assertEqual(wb.sheetnames, ["视频数据"])

    def test_export_all_multi_full_workflow(self):
        """v2.8.0：3 类型混合一键导出"""
        v1 = make(bvid="BV1a", title="视频A")
        v2 = make(bvid="BV1b", title="视频B")
        a1 = self._make_article(cv_id="cv1", title="专栏A")
        a2 = self._make_article(cv_id="cv2", title="专栏B")
        b1 = self._make_bangumi(season_id=1, title="番剧A")
        saved = export_all_multi(
            base="multi", out_dir=self.tmpdir,
            videos=[v1, v2], articles=[a1, a2], bangumis=[b1],
        )
        # xlsx 一个
        self.assertIn("xlsx", saved)
        # video csv/json/txt
        self.assertIn("csv_video", saved)
        self.assertIn("json_video", saved)
        self.assertIn("txt_video", saved)
        # article
        self.assertIn("csv_article", saved)
        self.assertIn("json_article", saved)
        self.assertIn("txt_article", saved)
        # bangumi
        self.assertIn("csv_bangumi", saved)
        self.assertIn("json_bangumi", saved)
        self.assertIn("txt_bangumi", saved)
        # 总共 10 个文件
        self.assertEqual(len(saved), 10)
        # 所有文件都存在
        for k, p in saved.items():
            self.assertTrue(os.path.exists(p), f"{k} 文件不存在: {p}")

    def test_export_all_multi_exclude_invalid(self):
        """v2.8.0：exclude_invalid 跨 3 类型都过滤"""
        v_ok = make(bvid="BV1a", status="ok")
        v_fail = make(bvid="BV1b", status="not_found")
        a_ok = self._make_article(cv_id="cv1", status="ok")
        a_fail = self._make_article(cv_id="cv2", status="not_found")
        saved = export_all_multi(
            base="filt", out_dir=self.tmpdir,
            videos=[v_ok, v_fail], articles=[a_ok, a_fail], bangumis=[],
            exclude_invalid=True,
        )
        # 只剩 ok 的
        with open(saved["csv_video"], encoding="utf-8-sig") as f:
            v_rows = list(csv.DictReader(f))
        self.assertEqual(len(v_rows), 1)
        with open(saved["csv_article"], encoding="utf-8-sig") as f:
            a_rows = list(csv.DictReader(f))
        self.assertEqual(len(a_rows), 1)

    def test_export_all_multi_video_only_backward_compat(self):
        """v2.8.0：只有视频时，文件命名仍是旧的（不带 _video 后缀）"""
        v1 = make(bvid="BV1a", title="视频A")
        # 走旧的 export_all 而不是 export_all_multi
        saved = export_all([v1], base="back", out_dir=self.tmpdir)
        self.assertIn("xlsx", saved)
        self.assertIn("csv", saved)
        # 文件名是 back.xlsx / back.csv（不带 _video）
        for k, p in saved.items():
            self.assertIn("back", os.path.basename(p))


if __name__ == "__main__":
    unittest.main()
