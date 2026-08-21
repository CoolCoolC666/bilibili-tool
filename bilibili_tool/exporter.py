"""导出器：xlsx / csv / json / txt 四种格式。

xlsx 走纯 openpyxl（不依赖 pandas），让脚本在精简 Python 环境也能跑。

**v2.8.0 扩展**：支持视频 / 专栏 / 番剧 三种 record 类型。
  - xlsx：3 个 sheet（视频/专栏/番剧）共用一个文件
  - csv/json/txt：3 个独立文件（{base}_video.csv / {base}_article.csv / {base}_bangumi.csv）
"""
from __future__ import annotations

import csv
import json
import os
from datetime import datetime
from typing import Any, Iterable, List, Tuple

from .models import ArticleInfo, BangumiInfo, VideoInfo


# ---- 视频 ----

# xlsx 顺序列
XLSX_COLUMNS = [
    ("raw_input", "原始输入"),
    ("input_kind", "输入类型"),
    ("status", "状态"),
    ("api_code", "API 状态码"),
    ("error", "错误信息"),
    ("aid", "AV号"),
    ("bvid", "BV号"),
    ("title", "标题"),
    ("up_name", "UP主"),
    ("tname", "分区"),
    ("view", "播放量"),
    ("like", "点赞"),
    ("coin", "投币"),
    ("favorite", "收藏"),
    ("share", "分享"),
    ("reply", "评论"),
    ("danmaku", "弹幕"),
    ("duration", "时长(秒)"),
    ("pubdate", "发布时间"),
    ("ctime", "投稿时间"),
    ("desc", "简介"),
    ("url", "链接"),
    ("fetched_at", "抓取时间"),
]

# ---- 专栏（v2.8.0 新增） ----

XLSX_COLUMNS_ARTICLE = [
    ("raw_input", "原始输入"),
    ("input_kind", "输入类型"),
    ("status", "状态"),
    ("api_code", "API 状态码"),
    ("error", "错误信息"),
    ("cv_id", "专栏ID(cv)"),
    ("title", "标题"),
    ("author_name", "作者"),
    ("author_mid", "作者UID"),
    ("view", "阅读量"),
    ("like", "点赞"),
    ("coin", "投币"),
    ("favorite", "收藏"),
    ("share", "分享"),
    ("reply", "评论"),
    ("words", "字数"),
    ("ctime", "创建时间"),
    ("pubtime", "发布时间"),
    ("summary", "摘要"),
    ("banner", "封面图URL"),
    ("url", "链接"),
    ("fetched_at", "抓取时间"),
]

# ---- 番剧（v2.8.0 新增） ----

XLSX_COLUMNS_BANGUMI = [
    ("raw_input", "原始输入"),
    ("input_kind", "输入类型"),
    ("status", "状态"),
    ("api_code", "API 状态码"),
    ("error", "错误信息"),
    ("season_id", "番剧季ID(ss)"),
    ("ep_id", "单集ID(ep)"),
    ("title", "标题"),
    ("alias", "别名"),
    ("type_name", "类型"),
    ("rating_score", "评分"),
    ("rating_count", "评分人数"),
    ("total_ep", "总集数"),
    ("status_text", "状态(已完结/连载中)"),
    ("publish_date", "上线日期"),
    ("view", "播放量"),
    ("favorite", "追番"),
    ("coins", "投币"),
    ("like", "点赞"),
    ("share", "分享"),
    ("reply", "评论"),
    ("danmaku", "弹幕"),
    ("desc", "简介"),
    ("cover", "封面图URL"),
    ("url", "链接"),
    ("fetched_at", "抓取时间"),
]


def _row_for_xlsx(v: Any, columns: list) -> dict:
    """根据 columns 列表，从 record 取出对应字段。"""
    d = v.to_dict() if hasattr(v, "to_dict") else dict(v.__dict__)
    return {label: d.get(key, "") for key, label in columns}


# ---- 去重（视频，按 bv/av） ----

def _dedupe_key(v: Any) -> Tuple:
    """按 record 自身字段去重。

    VideoInfo:     bv / av / raw
    ArticleInfo:   cv / raw
    BangumiInfo:   ss / ep / raw
    """
    if isinstance(v, VideoInfo):
        if v.bvid:
            return ("bv", v.bvid)
        if v.aid:
            return ("av", v.aid)
    elif isinstance(v, ArticleInfo):
        if v.cv_id:
            return ("cv", v.cv_id)
    elif isinstance(v, BangumiInfo):
        if v.season_id:
            return ("ss", v.season_id)
        if v.ep_id:
            return ("ep", v.ep_id)
    if v.raw_input:
        return ("raw", v.raw_input)
    return None


def dedupe_videos(videos: Iterable[VideoInfo]) -> List[VideoInfo]:
    """按 (bvid, aid, raw_input) 去重，**保留第一次出现**的那条。**保留旧 API 兼容**。"""
    seen: set = set()
    out: List[VideoInfo] = []
    for v in videos:
        key = _dedupe_key(v)
        if key is None:
            out.append(v)
            continue
        if key in seen:
            continue
        seen.add(key)
        out.append(v)
    return out


def dedupe_articles(articles: Iterable[ArticleInfo]) -> List[ArticleInfo]:
    """按 (cv_id, raw_input) 去重。"""
    seen: set = set()
    out: List[ArticleInfo] = []
    for a in articles:
        key = _dedupe_key(a)
        if key is None:
            out.append(a)
            continue
        if key in seen:
            continue
        seen.add(key)
        out.append(a)
    return out


def dedupe_bangumis(bangumis: Iterable[BangumiInfo]) -> List[BangumiInfo]:
    """按 (season_id, ep_id, raw_input) 去重。"""
    seen: set = set()
    out: List[BangumiInfo] = []
    for b in bangumis:
        key = _dedupe_key(b)
        if key is None:
            out.append(b)
            continue
        if key in seen:
            continue
        seen.add(key)
        out.append(b)
    return out


# ---- 过滤（按 status == "ok"） ----

def filter_valid(items: Iterable[Any]) -> List[Any]:
    """只保留 status == "ok" 的成功抓取记录。**保留旧 API 兼容**。"""
    return [v for v in items if getattr(v, "status", "") == "ok"]


# ---- 视频保存函数（保持旧 API 兼容） ----

def save_xlsx(videos: Iterable[VideoInfo], path: str) -> str:
    """单 sheet 视频导出（保留旧 API 兼容）。**新功能请用 save_xlsx_multi**。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "视频数据"

    headers = [label for _, label in XLSX_COLUMNS]
    ws.append(headers)

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3F6EE3")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    videos_list = list(videos)
    for v in videos_list:
        row = _row_for_xlsx(v, XLSX_COLUMNS)
        ws.append([row[label] for _, label in XLSX_COLUMNS])

    _set_column_widths(ws, videos_list, XLSX_COLUMNS)
    ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)
    return os.path.abspath(path)


def save_csv(videos: Iterable[VideoInfo], path: str) -> str:
    videos_list = list(videos)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[label for _, label in XLSX_COLUMNS],
        )
        writer.writeheader()
        for v in videos_list:
            row = _row_for_xlsx(v, XLSX_COLUMNS)
            writer.writerow(row)
    return os.path.abspath(path)


def save_json(videos: Iterable[VideoInfo], path: str) -> str:
    payload = [v.to_dict() for v in videos]
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return os.path.abspath(path)


def save_txt(videos: Iterable[VideoInfo], path: str) -> str:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    border = "═" * 60
    lines: List[str] = []
    lines.append("╔" + border + "╗")
    lines.append("║" + "🌟 B站视频信息提取清单 🌟".center(60) + "║")
    lines.append("╠" + border + "╣")
    lines.append(
        f"║ 🕒 导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}".ljust(62) + "║"
    )

    videos_list = list(videos)
    n_total = len(videos_list)
    lines.append(f"║ 📊 共 {n_total} 条记录".ljust(62) + "║")
    lines.append("╚" + border + "╝\n")

    for idx, v in enumerate(videos_list, 1):
        marker = "✓" if v.is_ok else "✗"
        title_disp = v.title or "(无标题)"
        lines.append(f"▶ [{idx:03d}] {marker} {title_disp}")
        lines.append(f"  👤 UP主: {v.up_name or '-'}")
        if v.tname:
            lines.append(f"  🏷️  分区: {v.tname}")
        if v.view is not None:
            lines.append(f"  👁  播放: {v.view:,} | 👍{v.like or 0:,} | 💬{v.reply or 0:,}")
        if v.pubdate:
            lines.append(f"  📅 发布时间: {v.pubdate}")
        if v.url:
            lines.append(f"  🔗 链接: {v.url}")
        if v.status != "ok":
            lines.append(f"  ⚠️  状态: {v.status} | code={v.api_code} | {v.error}")
        lines.append("  " + "-" * 56 + "\n")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return os.path.abspath(path)


# ---- 通用保存函数（v2.8.0） ----

def _set_column_widths(ws, items: list, columns: list) -> None:
    """根据内容动态设置列宽。"""
    from openpyxl.utils import get_column_letter
    for col_idx, (key, label) in enumerate(columns, 1):
        max_len = len(str(label))
        for v in items:
            d = v.to_dict() if hasattr(v, "to_dict") else dict(v.__dict__)
            val = d.get(key, "")
            s = "" if val is None else str(val)
            if len(s) > max_len:
                max_len = len(s)
        # 长文本列做宽一些
        if key in ("title", "desc", "summary", "evaluate", "url"):
            max_len = min(max_len, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)


def save_articles_csv(articles: Iterable[ArticleInfo], path: str) -> str:
    items = list(articles)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f, fieldnames=[label for _, label in XLSX_COLUMNS_ARTICLE]
        )
        writer.writeheader()
        for a in items:
            writer.writerow(_row_for_xlsx(a, XLSX_COLUMNS_ARTICLE))
    return os.path.abspath(path)


def save_articles_json(articles: Iterable[ArticleInfo], path: str) -> str:
    payload = [a.to_dict() for a in articles]
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return os.path.abspath(path)


def save_articles_txt(articles: Iterable[ArticleInfo], path: str) -> str:
    items = list(articles)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    border = "═" * 60
    lines: List[str] = []
    lines.append("╔" + border + "╗")
    lines.append("║" + "📝 B站专栏信息提取清单 📝".center(60) + "║")
    lines.append("╠" + border + "╣")
    lines.append(
        f"║ 🕒 导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}".ljust(62) + "║"
    )
    lines.append(f"║ 📊 共 {len(items)} 条记录".ljust(62) + "║")
    lines.append("╚" + border + "╝\n")
    for idx, a in enumerate(items, 1):
        marker = "✓" if a.is_ok else "✗"
        title_disp = a.title or "(无标题)"
        lines.append(f"▶ [{idx:03d}] {marker} {title_disp}")
        lines.append(f"  👤 作者: {a.author_name or '-'}")
        if a.words:
            lines.append(f"  📏 字数: {a.words}")
        if a.view is not None:
            lines.append(f"  👁  阅读: {a.view:,} | 👍{a.like or 0:,} | 💬{a.reply or 0:,}")
        if a.pubtime:
            lines.append(f"  📅 发布时间: {a.pubtime}")
        if a.url:
            lines.append(f"  🔗 链接: {a.url}")
        if a.summary:
            lines.append(f"  📋 摘要: {a.summary[:100]}{'...' if len(a.summary) > 100 else ''}")
        if a.status != "ok":
            lines.append(f"  ⚠️  状态: {a.status} | code={a.api_code} | {a.error}")
        lines.append("  " + "-" * 56 + "\n")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return os.path.abspath(path)


def save_bangumis_csv(bangumis: Iterable[BangumiInfo], path: str) -> str:
    items = list(bangumis)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f, fieldnames=[label for _, label in XLSX_COLUMNS_BANGUMI]
        )
        writer.writeheader()
        for b in items:
            writer.writerow(_row_for_xlsx(b, XLSX_COLUMNS_BANGUMI))
    return os.path.abspath(path)


def save_bangumis_json(bangumis: Iterable[BangumiInfo], path: str) -> str:
    payload = [b.to_dict() for b in bangumis]
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return os.path.abspath(path)


def save_bangumis_txt(bangumis: Iterable[BangumiInfo], path: str) -> str:
    items = list(bangumis)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    border = "═" * 60
    lines: List[str] = []
    lines.append("╔" + border + "╗")
    lines.append("║" + "🎬 B站番剧信息提取清单 🎬".center(60) + "║")
    lines.append("╠" + border + "╣")
    lines.append(
        f"║ 🕒 导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}".ljust(62) + "║"
    )
    lines.append(f"║ 📊 共 {len(items)} 条记录".ljust(62) + "║")
    lines.append("╚" + border + "╝\n")
    for idx, b in enumerate(items, 1):
        marker = "✓" if b.is_ok else "✗"
        title_disp = b.title or "(无标题)"
        lines.append(f"▶ [{idx:03d}] {marker} {title_disp}")
        if b.alias:
            lines.append(f"  🏷️  别名: {b.alias}")
        if b.type_name:
            lines.append(f"  🎬 类型: {b.type_name}")
        if b.rating_score:
            lines.append(f"  ⭐ 评分: {b.rating_score} ({b.rating_count or 0:,}人)")
        if b.total_ep:
            lines.append(f"  📺 总集数: {b.total_ep}")
        if b.status_text:
            lines.append(f"  🟢 状态: {b.status_text}")
        if b.view is not None:
            lines.append(f"  👁  播放: {b.view:,} | 👍{b.like or 0:,} | 💬{b.reply or 0:,}")
        if b.publish_date:
            lines.append(f"  📅 上线: {b.publish_date}")
        if b.url:
            lines.append(f"  🔗 链接: {b.url}")
        if b.status != "ok":
            lines.append(f"  ⚠️  状态: {b.status} | code={b.api_code} | {b.error}")
        lines.append("  " + "-" * 56 + "\n")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return os.path.abspath(path)


def save_xlsx_multi(
    videos: List[VideoInfo],
    articles: List[ArticleInfo],
    bangumis: List[BangumiInfo],
    path: str,
) -> str:
    """v2.8.0：xlsx 3 个 sheet 共用一个文件。每个 sheet 可为空。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # 默认 sheet 删掉，重新建
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3F6EE3")
    alignment = Alignment(horizontal="center", vertical="center")

    def add_sheet(name: str, items: list, columns: list, color: str = "3F6EE3") -> None:
        if not items:
            return  # 跳过空数据
        ws = wb.create_sheet(name)
        headers = [label for _, label in columns]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = alignment
        for v in items:
            row = _row_for_xlsx(v, columns)
            ws.append([row[label] for _, label in columns])
        _set_column_widths(ws, items, columns)
        ws.freeze_panes = "A2"

    add_sheet("视频数据", videos, XLSX_COLUMNS, "3F6EE3")
    add_sheet("专栏数据", articles, XLSX_COLUMNS_ARTICLE, "E67E22")
    add_sheet("番剧数据", bangumis, XLSX_COLUMNS_BANGUMI, "27AE60")

    # 如果三个 sheet 全空，至少建一个空 sheet
    if not wb.sheetnames:
        ws = wb.create_sheet("无数据")
        ws["A1"] = "本次抓取无任何数据"

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)
    return os.path.abspath(path)


# ---- 一键导出（v2.8.0 旧 API 兼容） ----

def export_all(
    videos: List[VideoInfo],
    *,
    base: str,
    out_dir: str = "output",
    dedupe: bool = True,
    exclude_invalid: bool = False,
) -> dict:
    """视频版 export_all（保留旧 API 兼容）。**新功能请用 export_all_multi**。"""
    if exclude_invalid:
        videos = filter_valid(videos)
    if dedupe:
        videos = dedupe_videos(videos)
    saved = {}
    saved["xlsx"] = save_xlsx(videos, os.path.join(out_dir, f"{base}.xlsx"))
    saved["csv"] = save_csv(videos, os.path.join(out_dir, f"{base}.csv"))
    saved["json"] = save_json(videos, os.path.join(out_dir, f"{base}.json"))
    saved["txt"] = save_txt(videos, os.path.join(out_dir, f"{base}.txt"))
    return saved


# ---- 一键导出（v2.8.0 新增：3 类型混合） ----

def export_all_multi(
    *,
    base: str,
    out_dir: str = "output",
    videos: List[VideoInfo] = None,
    articles: List[ArticleInfo] = None,
    bangumis: List[BangumiInfo] = None,
    dedupe: bool = True,
    exclude_invalid: bool = False,
) -> dict:
    """v2.8.0：3 类型混合导出。

    xlsx  = {base}.xlsx                    （3 个 sheet）
    csv   = {base}_video.csv / _article.csv / _bangumi.csv
    json  = {base}_video.json / _article.json / _bangumi.json
    txt   = {base}_video.txt / _article.txt / _bangumi.txt

    每个类型可选：传入空 list 跳过该类型。
    处理顺序：先按 exclude_invalid 过滤 → 再按 dedupe 去重。
    """
    videos = list(videos or [])
    articles = list(articles or [])
    bangumis = list(bangumis or [])

    if exclude_invalid:
        videos = filter_valid(videos)
        articles = filter_valid(articles)
        bangumis = filter_valid(bangumis)
    if dedupe:
        videos = dedupe_videos(videos)
        articles = dedupe_articles(articles)
        bangumis = dedupe_bangumis(bangumis)

    saved = {}
    # xlsx 一个文件，3 sheet
    saved["xlsx"] = save_xlsx_multi(videos, articles, bangumis, os.path.join(out_dir, f"{base}.xlsx"))
    # csv / json / txt 各 3 个文件
    if videos:
        saved["csv_video"] = save_csv(videos, os.path.join(out_dir, f"{base}_video.csv"))
        saved["json_video"] = save_json(videos, os.path.join(out_dir, f"{base}_video.json"))
        saved["txt_video"] = save_txt(videos, os.path.join(out_dir, f"{base}_video.txt"))
    if articles:
        saved["csv_article"] = save_articles_csv(articles, os.path.join(out_dir, f"{base}_article.csv"))
        saved["json_article"] = save_articles_json(articles, os.path.join(out_dir, f"{base}_article.json"))
        saved["txt_article"] = save_articles_txt(articles, os.path.join(out_dir, f"{base}_article.txt"))
    if bangumis:
        saved["csv_bangumi"] = save_bangumis_csv(bangumis, os.path.join(out_dir, f"{base}_bangumi.csv"))
        saved["json_bangumi"] = save_bangumis_json(bangumis, os.path.join(out_dir, f"{base}_bangumi.json"))
        saved["txt_bangumi"] = save_bangumis_txt(bangumis, os.path.join(out_dir, f"{base}_bangumi.txt"))
    return saved
